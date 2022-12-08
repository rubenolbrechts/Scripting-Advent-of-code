#!/usr/bin/python3
################################################################################
# Excel files (example-code-5-xlsx.py)
################################################################################
# sudo dnf install python3-openpyxl
from openpyxl import Workbook
################################################################################
# Initialize Workbook
wb = Workbook()
# Active worksheet
ws = wb.active
# Assign data directly to cell
# in A column on first row
ws['A1'] = 99
# Append row
ws.append([1, 2, 3])
# Python types automatically converted
from datetime import date
ws['A2'] = date.today()
################################################################################
# Set worksheet name
ws.title = "My first worksheet"
for row in range(3, 10):
    ws.append(range(5))
# Create second worksheet
ws2 = wb.create_sheet(title="WS2")
ws2['B2'] = "This is B2"
for row in range(10, 21):
    for col in range(1, 11):
        cell_value = "c" + str(col) + "-" + "r" + str(row)
        #print
        _ = ws2.cell(column=col, row=row, value="{0}".format(cell_value))
# Save file
wb.save("example.xlsx")
################################################################################
# Read existing workbook
from openpyxl import load_workbook
wb2read = load_workbook(filename = 'example.xlsx')
print("Reading from workbook example.xlsx worksheet WS2:")
wb2read_ws2 = wb2read['WS2']
print(wb2read_ws2['F16'].value)
################################################################################
# Merge cells
wb2 = Workbook()
wb2_ws1 = wb2.active
wb2_ws1.merge_cells('A2:D2')
# .unmerge_cells('A2:D2')
wb2_ws1['A2'] = "This cell is merged"
wb2_ws1.merge_cells(start_row=3, start_column=1, end_row=5, end_column=4)
# .unmerge_cells(start_row=3, start_column=1, end_row=5, end_column=4)
wb2_ws1['A3'] = "Even more\ncells merged"
wb2.save("example2.xlsx")
################################################################################
# Cells with styles
wb3 = Workbook()
ws = wb3.active
from openpyxl.styles import Font
# Example 1: enter value cell and change style 
ws['C3'] = "C3 has style"
c = ws['C3']
c.font = Font(name='Arial', size=20, color="FF0000")
# Example 2: set style, use style for cell
style1 = Font(name='Tahoma', size=16, color="0000FF", bold=True)
ws['A5'] = "A5"
a5 = ws['A5']
b5 = ws['B5']
a5.font = style1
b5.font = style1
ws['B5'] = "B5"
# Example 3: with borders and aligned
from openpyxl.styles import Side, Border, Alignment
d10 = ws['D10']
d10.value = "borders"
d10.font = Font(name='Tahoma', size=10, color="600060", bold=True)
thin = Side(border_style="thin", color="000000")
double = Side(border_style="double", color="ff0000")
d10.border = Border(top=double, left=thin, right=thin, bottom=double)
d10.alignment = Alignment(horizontal="center", vertical="center")
# after written to cell in column 
# you can modify column dimensions
ws.column_dimensions["C"].width = 30.0
ws.row_dimensions[10].height = 40.0
# Save workbook
wb3.save('workbook_with_style.xlsx')
################################################################################